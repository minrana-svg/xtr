import streamlit as st
import camelot
import pandas as pd
import re
import tempfile
import os
from io import BytesIO
from img2table.document import PDF as Img2TablePDF
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import cv2
import numpy as np
from typing import Optional

# Monkey patch for img2table to avoid niBlackThreshold issue
def threshold_dark_areas(img: np.ndarray, char_length: Optional[float]) -> np.ndarray:
    """
    Threshold image by differentiating areas with light and dark backgrounds
    :param img: image array
    :param char_length: average character length
    :return: threshold image
    """
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    # If image is mainly black, revert the image
    if np.mean(gray) <= 127:
        gray = 255 - gray

    thresh_kernel = int(char_length) // 2 * 2 + 1
    if thresh_kernel % 2 == 0:
        thresh_kernel += 1  # Make sure it's odd

    # Use adaptive threshold instead of niBlackThreshold
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, thresh_kernel, 2)
    binary_thresh = None

    # Mask on areas with dark background
    blur_size = min(255, int(2 * char_length) // 2 * 2 + 1)
    if blur_size % 2 == 0:
        blur_size += 1
    blur = cv2.GaussianBlur(gray, (blur_size, blur_size), 0)
    mask = cv2.inRange(blur, 0, 100)

    # Identify dark areas
    _, _, stats, _ = cv2.connectedComponentsWithStats(mask, 8, cv2.CV_32S)

    # For each dark area, use binary threshold instead of regular threshold
    for idx, row in enumerate(stats):
        # Get statistics
        x, y, w, h, area = row

        if idx == 0:
            continue

        if area / (w * h) >= 0.5 and min(w, h) >= char_length and max(w, h) >= 5 * char_length:
            if binary_thresh is None:
                # Use adaptive threshold for binary image
                bin_thresh = cv2.adaptiveThreshold(255 - gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, thresh_kernel, 2)
                binary_thresh = 255 - bin_thresh  # Invert if needed
            thresh[y:y+h, x:x+w] = binary_thresh[y:y+h, x:x+w]

    return thresh

# Apply monkey patch
import img2table.tables
img2table.tables.threshold_dark_areas = threshold_dark_areas

# --- 内部処理用の関数 ---
def parse_pages(s):
    r = []
    for p in s.split(","):
        if "-" in p:
            start, end = map(int, p.split("-"))
            r.extend(range(start, end + 1))
        else:
            r.append(int(p))
    return r

def chunk_pages(p, s=5):
    return [p[i:i+s] for i in range(0, len(p), s)]

def filter_img2table(df):
    return not(df is None or df.empty or df.shape[0]<3 or df.shape[1]<2) and \
           df.replace("", pd.NA).notna().sum().sum() / (df.shape[0]*df.shape[1]) >= 0.3

def calc_stats(df):
    t = df.shape[0] * df.shape[1]
    f = df.replace("", pd.NA).notna().sum().sum()
    return f, t

def extract_all_tables(pdf_path, pages, line_scale=40):
    r = {}
    for p in pages:
        pt = []
        try:
            # line_scaleは常にデフォルトの40を使用
            l = camelot.read_pdf(str(pdf_path), pages=str(p), flavor="lattice", line_scale=line_scale)
            for t in l:
                if hasattr(t, "df"): pt.append(("lattice", t.df))
        except Exception: pass
        
        try:
            s = camelot.read_pdf(str(pdf_path), pages=str(p), flavor="stream")
            for t in s:
                if hasattr(t, "df"): pt.append(("stream", t.df))
        except Exception: pass
        
        if pt: r[p] = pt

    for ch in chunk_pages(pages, 5):
        try:
            pdf = Img2TablePDF(src=str(pdf_path), pages=[x-1 for x in ch], pdf_text_extraction=True)
            ex = pdf.extract_tables(implicit_rows=True, borderless_tables=True, min_confidence=50)
            for p in ch:
                for t in ex.get(p-1, []):
                    try:
                        if hasattr(t, "df") and filter_img2table(t.df):
                            r.setdefault(p, []).append(("img2table", t.df))
                    except Exception as e:
                        st.warning(f"img2tableスキップ {p}: {e}")
        except Exception as e:
            st.warning(f"chunk失敗 {ch}: {e}")
    return r

def export_to_excel(all_results, pages, mode):
    output = BytesIO() # メモリ上にExcelを保存する
    with pd.ExcelWriter(output, engine="openpyxl") as w:
        wb = w.book
        if mode == "separate":
            for fn, res in all_results:
                sh = fn[:31] # シート名の文字数制限対策
                sr = 0
                for p in pages:
                    tb = res.get(p, [])
                    if not tb: continue
                    pd.DataFrame([[f"{p}ページ"]]).to_excel(w, sheet_name=sh, startrow=sr, index=False, header=False)
                    sr += 2
                    for src, df in tb:
                        f, t = calc_stats(df)
                        pd.DataFrame([[f"[{src}] セル数:{f}/{t}"]]).to_excel(w, sheet_name=sh, startrow=sr, index=False, header=False)
                        sr += 1
                        df.to_excel(w, sheet_name=sh, startrow=sr, index=False, header=False)
                        sr += len(df) + 3
        else:
            sh = "まとめ"
            cc = 0
            for fn, res in all_results:
                sr = 1
                pd.DataFrame([[fn]]).to_excel(w, sheet_name=sh, startrow=0, startcol=cc, index=False, header=False)
                mw = 0
                for p in pages:
                    tb = res.get(p, [])
                    if not tb: continue
                    pd.DataFrame([[f"{p}ページ"]]).to_excel(w, sheet_name=sh, startrow=sr, startcol=cc, index=False, header=False)
                    sr += 2
                    for src, df in tb:
                        f, t = calc_stats(df)
                        pd.DataFrame([[f"[{src}] セル数:{f}/{t}"]]).to_excel(w, sheet_name=sh, startrow=sr, startcol=cc, index=False, header=False)
                        sr += 1
                        df.to_excel(w, sheet_name=sh, startrow=sr, startcol=cc, index=False, header=False)
                        mw = max(mw, df.shape[1])
                        sr += len(df) + 3
                cc += mw + 3
                
        # 見た目のフォーマット調整
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r_cells in ws.iter_rows():
                for c in r_cells:
                    if not c.value: continue
                    t_val = str(c.value)
                    if c.row == 1 or re.match(r"^\d+ページ$", t_val):
                        c.font = Font(bold=True, color="FF0000", size=14)
                    elif "[" in t_val:
                        c.font = Font(bold=True, size=12)
            for col in ws.columns:
                ml = 0
                cl = get_column_letter(col[0].column)
                for c in col:
                    if c.value: ml = max(ml, len(str(c.value)))
                ws.column_dimensions[cl].width = min(ml + 2, 50)
                
    output.seek(0)
    return output

# --- Web UI (Streamlit) ---
st.set_page_config(page_title="PDF表抽出ツール", layout="wide")
st.title("📄 PDF表抽出ツール")

# 入力フォーム (line_scaleを削除し、ページ指定のみに変更)
st.sidebar.header("設定")
pages_str = st.sidebar.text_input("ページ数指定 (例: 7-12, 14)", value="7-12")

uploaded_files = st.file_uploader("PDFファイルをアップロードしてください（複数可）", type="pdf", accept_multiple_files=True)

if st.button("表を抽出してExcelを作成", type="primary"):
    if not uploaded_files:
        st.error("⚠️ PDFファイルがアップロードされていません。")
    else:
        pages = parse_pages(pages_str)
        all_results = []
        
        # 進行状況バー
        progress_bar = st.progress(0)
        
        with tempfile.TemporaryDirectory() as tmpdir:
            for i, uploaded_file in enumerate(uploaded_files):
                st.write(f"⏳ 処理中: {uploaded_file.name}")
                temp_path = os.path.join(tmpdir, uploaded_file.name)
                
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                try:
                    # 裏側でline_scale=40として処理を実行
                    res = extract_all_tables(temp_path, pages, line_scale=40)
                    stem = os.path.splitext(uploaded_file.name)[0]
                    all_results.append((stem, res))
                except Exception as e:
                    st.error(f"❌ {uploaded_file.name} でエラー: {e}")
                
                progress_bar.progress((i + 1) / len(uploaded_files))
        
        if all_results:
            st.success("✅ 抽出が完了しました！下のボタンからダウンロードできます。")
            
            # Excelファイルの生成
            excel_separate = export_to_excel(all_results, pages, "separate")
            excel_single = export_to_excel(all_results, pages, "single")
            
            # ダウンロードボタンの配置
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 統合結果_separate.xlsx をダウンロード",
                    data=excel_separate,
                    file_name="統合結果_separate.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col2:
                st.download_button(
                    label="📥 統合結果_single.xlsx をダウンロード",
                    data=excel_single,
                    file_name="統合結果_single.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
